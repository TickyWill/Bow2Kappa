'''
Definition of useful functions for bow and cutting efficiency plots in batch
 - Bow_Kappa_files
 - Param_extract
 - Data_extract
 - bow2kappa
 - Kappa_3Dplot_values
 - Kappa_Wireframe
 '''

__all__ = [
    "Bow_Kappa_files",
    "Param_extract",
    "Data_extract",
    "bow2kappa",
    "Kappa_3Dplot_values",
    "Kappa_Wireframe",
]

# Global variables used by bow2kappa function
LONG = 121      # Length of the moving window for bow data smoothing; must be odd
WIN = 'boxcar'  # Kind of moving window; 
                # in ['boxcar', 'triang', 'hamming', 'blackman', 'bartlett', 'parzen', 'bohman'] 

def Bow_Kappa_files(cut_name, my_path):
    '''
    
    Set file names for a cut

    Args
    - cut_name (string): identification of the cut (ex: "125" where 125 is the reference number of the cut)

    - my_path (pathlib.PosixPath): path for input and output files

    Returns
    - input_file (pathlib.PosixPath): file name for reading the experimental data of the cut

    - output_file (pathlib.PosixPath): file name for saving the results related to the cut

    '''
    input_file_name = 'Bow_' + cut_name + '.xlsx'
    output_file_name = 'Results_' + cut_name + '.xlsx'
    input_file = my_path / input_file_name
    output_file = my_path / output_file_name

    return input_file,output_file


def Param_extract(file, sheet):
    '''
     
    Extract the cutting parameters.
    The parameters are listed in the sheet 'sheet' of the EXCEL file 'file'.

    Args
    - file (pathlib.PosixPath): file name for reading the experimental data of the cut

    - sheet (string): name of the sheet containing the parameters in the EXCEL file 'file'

    Returns
    - dparam (dictionary): from which cutting parameters values may be extracted using key "Value"

    '''

    # 3rd party imports
    import pandas as pd

    # Get the values of the cutting parameters from the input file
    dparam = pd.read_excel(file,
                         sheet_name = sheet)

    return dparam

def Data_extract(file, sheet, sensors_nbr):
    '''
     
    Extract the data listed in the sheet 'sheet' of the EXCEL file 'file'.

    Args
    - file (pathlib.PosixPath): file name for reading the  data of the cut

    - sheet (string): name of the sheet containing the data in the EXCEL file 'file'

    - sensors_nbr (integer): number of sensors of which the data will be extracted

    Returns
    - data (dictionary): data extracted from the EXCEL file 'file' using columns C up to G
    depending on the number of sensors sensors_nbr vs cut progress (column B) expressed either in time
    or in percentage of cut duration

    '''

    # 3rd party imports
    import pandas as pd

    # Get the bow data from the input file
    kappa_cols = {
                 0: "B",
                 1: "C",
                 2: "D",
                 3: "E",
                 4: "F",
                 5: "G"
                 }
    endcol=kappa_cols[sensors_nbr]
    usecols = 'B:'+ str(endcol)

    data = pd.read_excel(file,
                         sheet_name =sheet,
                         usecols=usecols)

    return data

def bow2kappa(data, dparam, output_file):

    '''
      
    Data filering and downsampling
    Vertical force computation
    Cutting efficiency kappa computation
    Storage of the results in an excel file with three sheets:
         Bow with columns names:  'Cut progress (%)' 'Bow # (mm)'
         Force with columns names: 'Cut progress (%)' 'Force # (N)'
         Kappa with columns names: 'Cut progress (%)' 'Kappa # x10^7 (m/N)'

    Args
    - data (dataframe): bow raw measurements

    - dparam (dataframe): from which cutting parameters values may be extracted using key "Value"

    - output_file (pathlib.PosixPath): file name for saving the results of the computations

    Returns
    - dic_bow (dataframe): bow data after smoothing and downsampling

    - dic_force (dataframe): force data as calculated using dic_bow

    - dic_kappa (dataframe): kappa data as calculated using dic_force

    '''

    # Standard library imports
    import os

    # 3rd party imports
    import numpy as np
    import pandas as pd
    from scipy import signal
    import openpyxl
    import matplotlib.pyplot as plt

    # Excel writer to flush the results
    try:
        with open(output_file):
            os.remove(output_file)
            writer = pd.ExcelWriter(output_file, engine='openpyxl')
    except IOError:
        writer = pd.ExcelWriter(output_file, engine='openpyxl')

    # Get useful parameters for bow conversion to cutting efficiency
    time_to_contact = np.array(dparam['Value'])[2]
    cut_effective_duration = np.array(dparam['Value'])[3]
    sensors_nbr = np.array(dparam['Value'])[4]
    sensor_init = np.array(dparam['Value'])[5]
    wire_guides_gap = np.array(dparam['Value'])[6]
    brick_width = np.array(dparam['Value'])[7]
    wire_tension = np.array(dparam['Value'])[8]
    table_speed = (np.array(dparam['Value'])[9])/(1000*60)
    wire_speed = np.array(dparam['Value'])[10]

    # conversion time --> cut percentage
    per_cent_coupe = np.array(data['Time (s)'])
    time_init = per_cent_coupe[0]
    per_cent_coupe = 100 * (per_cent_coupe - time_init - time_to_contact) / (60*cut_effective_duration)

    # Smooth and downsample the data using a moving window of length long and of type win
    long = LONG
    win = WIN 

    norme = sum(signal.get_window(win, long))
    dic_bow = {}
    bow_name = [x for x in data.columns if 'Bow' in x]
    for bow in bow_name:
        data_filter = data[bow].rolling(long,center=True, win_type=win, axis=0).sum()/norme
        dic_bow[bow] = data_filter[long//2+1::long]
    dic_bow['Cut progress (%)'] = per_cent_coupe[long//2+1::long]
    new = pd.DataFrame.from_dict(dic_bow)
    new=new.reindex(columns= ['Cut progress (%)', *bow_name])

    if os.path.exists(output_file):
            book = openpyxl.load_workbook(output_file)
            writer.book = book

    new.to_excel(writer, sheet_name='Bow')
    writer.save()
    writer.close()

    # Vertical force computation
    dic_force = {}
    force_name = [bow.replace('Bow','Force')[:-5] + ' (N)' for bow in bow_name] # Set force columns names with bow columns names
                                                                                # [:-5] suppress " (mm)" in bow columns names
    for force,bow in zip(force_name,bow_name):
        dic_force[force] = (4*wire_tension*np.array((dic_bow[bow]))) \
        /(2*wire_guides_gap+brick_width)
    dic_force['Cut progress (%)'] = per_cent_coupe[long//2+1::long]
    new_force = pd.DataFrame.from_dict(dic_force)
    new_force = new_force.reindex(columns= ['Cut progress (%)', *force_name])

    if os.path.exists(output_file):
            book = openpyxl.load_workbook(output_file)
            writer.book = book

    new_force.to_excel(writer, sheet_name='Force')
    writer.save()
    writer.close()

    # Cutting efficiency computation
    dic_kappa = {}
    kappa_name = [bow.replace('Bow','Kappa')[:-5] + ' x10^7 (m/N)' for bow in bow_name] # Set kappa columns names with bow columns names
                                                                                        # [:-5] suppress " (mm)" in bow columns names
    for kappa, force in zip(kappa_name,force_name):
        dic_kappa[kappa] = 10000000*(brick_width/1000)*table_speed \
        /(wire_speed*np.array((dic_force[force])))
    dic_kappa['Cut progress (%)'] = per_cent_coupe[long//2+1::long]
    new_kappa = pd.DataFrame.from_dict(dic_kappa)
    new_kappa = new_kappa.reindex(columns= ['Cut progress (%)', *kappa_name])

    if os.path.exists(output_file):
            book = openpyxl.load_workbook(output_file)
            writer.book = book

    new_kappa.to_excel(writer, sheet_name='Kappa')
    writer.save()
    writer.close()

    return dic_bow,dic_force,dic_kappa

def Kappa_3Dplot_values(dkappa,val_min, val_max,z_min,z_max,sensors_nbr):

    '''
     
    Arrangement of the cutting efficiency calculated from the bow in-situ measurements
    by bow2kappa function for a wireframe 3D plot

    Args
    - output_file: name of the results file (output of bow2kappa function)

    - val_min, val_max : minimum and maximum values of the cut progress range
      used for data selection and xaxis in the 3D plot

    - z_min, z_max: minimum and maximum values of the cutting efficiency range
      used for zaxis in the 3D plot

    - sensors_nbr: number of sensors used for the bow in-situ measurements (parameter of the cut)

    Return
    - 3D plot of the cutting efficiency values

    '''

    # 3rd party imports
    import numpy as np
    import pandas as pd
    from scipy import interpolate

    # Select cut progress range and build x,y,z values distribution suitable for the 3D plot
    dkappa.rename({"Cut progress (%)":"A"}, axis=1, inplace=True) # Rename colonne for query attribute use
    df= dkappa.query('A > @val_min and A < @val_max')
    u = [i for i in range(1,sensors_nbr+1)] # Sensor number
    v = list(df['A'])                       # % of cut
    cut_progress_nb= len(v)                 # Number of cut progress values
    df=df.drop(['A'],axis=1)                # Keep only z values in df
    z = df.values                           # 2D array of z values

    # z calculation at interpolated values of u,v
        # Definition of 2D linear interpolation function
    newfunc = interpolate.interp2d(u,v,z,kind='linear')
        # New sampling of 1000 points along x between extrema of x1
    unew = np.linspace(u[0] , u[-1], num=1000, endpoint=True)
        # New sampling of 1000 points along y between extrema of y1
    vnew = np.linspace(v[0] , v[-1], num=1000, endpoint=True)
        # Use of the defined newfunc as 2D linear interpolation
        # (x,y) mesh re-construction for the (xnew,ynew) positions
    y,x = np.meshgrid(unew, vnew)
        # for the calculation of z at the (xnew,ynew) positions
    znew = newfunc(unew, vnew)

    return x, y, znew, cut_progress_nb

def Kappa_Wireframe(x, y, z, z_min, z_max, cut_name, sensors_nbr, cut_progress_nb, ax, factor = 1):

    '''
    Wireframe 3D plot configuration of the cutting efficiency values
    as function of the cut progress and the sensor number

    Args
    - x:  1D array of the interpolated values of the cut progress for the wireframe 3D plot
      as arranged by Kappa_3Dplot_values function

    - y:  1D array of the interpolated values of the sensor number for the wireframe 3D plot
      as arranged by Kappa_3Dplot_values function

    - z: 2D array of cutting efficiency as arranged by Kappa_3Dplot_values function

    - z_min, z_max: minimum and maximum values of z range
      used for zaxis in the 3D plot

    - cut_name (string): identification of the cut (ex: "125" where 125 is the reference number of the cut)

    - sensors_nbr (integer): number of sensors used for the bow in-situ measurements (parameter of the cut)

    - cut_progress_nb (integer): number of cut progress sampling points (parameter of the cut)

    - ax (matplotlib.axes._subplots.AxesSubplot): figure subplot description

    Returns
    - configuration of z values as function of x and y for a wireframe 3D plot

    '''

    # 3rd party imports
    import matplotlib.pyplot as plt

    # Set plot parameters
    ticks_size = 10 * factor # fontsize of ticks labels
    label_size = 12 * factor # fontsize of axes labels
    line_space = 2 * factor  # line spacing between ticks labels and axe label
    x0 = [10,20,30,40,50,60,70,80]                  # xticklabels
    y0 = [i for i in range(1,sensors_nbr+1)]        # yticklabels
    z_min_rnd,z_max_rnd,z_step_nbr= int(round(z_min)),int(round(z_max)),1
    z0=[i for i in range(z_min_rnd+1,z_max_rnd+1)]  # yticklabels
    thickness = 0.4  # Line thickness of the 3D plot curves
    plot_azim = 40   # 3D plot azimut
    plot_elev = 5    # 3D plot elevation

    # Set a wireframe 3D plot
    surf=ax.plot_wireframe(x, y, z, thickness, rcount=cut_progress_nb, \
                           ccount=sensors_nbr-1, color='red')

    # Set plot title
    ax.set_title(f'Cut {cut_name}', fontsize = label_size)
    # Set point of view of 3D plot
    ax.view_init(elev=plot_elev, azim=plot_azim)

    # Set x axe
    ax.set_xticks(x0, minor=False)
    ax.set_xticklabels(x0, fontsize=ticks_size)
    ax.set_xlabel('\nCut progress (%)', linespacing=line_space, fontsize=label_size)

    # Set y axe
    ax.set_yticks(y0, minor=False)          # set ticks label for y axis (integer values)
    ax.set_yticklabels(y0, fontsize=ticks_size)
    ax.set_ylabel('\nSensor number', linespacing=line_space, fontsize=label_size)

    # Set z axe
    ax.set_zlim(z_min, z_max)   # set limits of z axis
    ax.set_zticks(z0, minor=False)
    ax.set_zticklabels(z0, fontsize=ticks_size)
    ax.zaxis.set_rotate_label(False)
    ax.set_zlabel('Cutting efficiency (10$^{-7}$ m.N$^{-1}$)', \
                  rotation=90, fontsize=label_size)
